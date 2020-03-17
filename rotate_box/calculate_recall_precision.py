# -*- coding: utf-8 -*-

import xml.etree.ElementTree as ET
import os
import sys
import glob
from PIL import Image
import openpyxl as px
from openpyxl.styles import Alignment, borders
from openpyxl.styles.fonts import Font
from termcolor import cprint
import numpy as np
import cv2
import time
from tqdm import tqdm
import argparse


parser = argparse.ArgumentParser(description = 'Caluclate IOU and Detection rate')

parser.add_argument('-c', '--category',      default = './category.txt')
parser.add_argument('-px', '--pre_xml',      default = './data/R2CNN/predict/labelxml_rotate0.6')
parser.add_argument('-cx', '--cor_xml',      default = './data/R2CNN/20190709_txt_clip_rotate/val/labelxml')
parser.add_argument('-i', '--iou_threshold', default = 0.6)

# action ---> オプション--✕✕をつけるとTrue、つけないとデフォルト値のFalse
parser.add_argument('-r', '--rotate_box',    action='store_true')

args = parser.parse_args()


"""
---------------------------------------------------------------------------------------------------

XMLファイルからラベルを取得

---------------------------------------------------------------------------------------------------
"""
def get_label(xml_list, categories, dirform=True, correct=False):
    label_list = []

    for xml_path in xml_list:
        tree = ET.parse(xml_path)

        framename = ''

        if (dirform):
            directory = xml_path.rsplit('/', 3)[1]
            framename = xml_path.rsplit('/', 3)[3].rsplit('.', 1)[0]
            framename = os.path.join(directory, framename)

        else:
            framename = xml_path.rsplit('/', 1)[1].rsplit('.', 1)[0]

        num = 0

        for object in tree.findall('object'):
            labelname = ''

            for cat in object.findall('category'):
                for cls in categories:
                    if (cls == cat.find('value').text):
                        labelname = cls

            if (object.find('bndbox').find('xmin') == None):
                sys.exit('Missmatch Box Type')

            xmin = int(object.find('bndbox').find('xmin').text)
            ymin = int(object.find('bndbox').find('ymin').text)
            xmax = int(object.find('bndbox').find('xmax').text)
            ymax = int(object.find('bndbox').find('ymax').text)

            if not (correct):
                score = object.find('score').text
                label_list.append([framename, labelname, xmin, ymin, xmax, ymax, num, score])

            else:
                label_list.append([framename, labelname, xmin, ymin, xmax, ymax, num])

            # １枚の画像に含まれるバウンディングボックス数をカウント
            num += 1

    # 昇順にソート
    label_list.sort()

    # label_list = [directory/framename or framename, labelname, x1, y1, x2, y2, num, (score)]
    return label_list


"""
---------------------------------------------------------------------------------------------------

XMLファイルからラベルを取得（回転ver.）

---------------------------------------------------------------------------------------------------
"""
def get_label_rotate(xml_list, categories, dirform=True, correct=False):
    label_list = []

    for xml_path in xml_list:
        tree = ET.parse(xml_path)

        framename = ''

        if (dirform):
            directory = xml_path.rsplit('/', 3)[1]
            framename = xml_path.rsplit('/', 3)[3].rsplit('.', 1)[0]
            framename = os.path.join(directory, framename)

        else:
            framename = xml_path.rsplit('/', 1)[1].rsplit('.', 1)[0]

        num = 0

        for object in tree.findall('object'):
            labelname = ''

            for cat in object.findall('category'):
                for cls in categories:
                    if (cls == cat.find('value').text):
                        labelname = cls

            if (object.find('bndbox').find('x0') == None):
                sys.exit('Missmatch Box Type')

            x0 = int(object.find('bndbox').find('x0').text)
            y0 = int(object.find('bndbox').find('y0').text)
            x1 = int(object.find('bndbox').find('x1').text)
            y1 = int(object.find('bndbox').find('y1').text)
            x2 = int(object.find('bndbox').find('x2').text)
            y2 = int(object.find('bndbox').find('y2').text)
            x3 = int(object.find('bndbox').find('x3').text)
            y3 = int(object.find('bndbox').find('y3').text)

            if not (correct):
                score = object.find('score').text
                label_list.append([framename, labelname, x0, y0, x1, y1, x2, y2, x3, y3, num, score])

            else:
                label_list.append([framename, labelname, x0, y0, x1, y1, x2, y2, x3, y3, num])

            # １枚の画像に含まれるバウンディングボックス数をカウント
            num += 1

    # 昇順にソート
    label_list.sort()

    # label_list = [directory/framename or framename, labelname, x1, y1, x2, y2, num, score]
    return label_list


"""
---------------------------------------------------------------------------------------------------

ボックス合計数を取得

---------------------------------------------------------------------------------------------------
"""
def get_boxnum(label_list, CLASSES):
    boxnum_list = []

    for category in CLASSES:
        cnt = [x[1] for x in label_list].count(category)

        boxnum_list.append([category, cnt])

    return boxnum_list


"""
---------------------------------------------------------------------------------------------------

IOUを計算

---------------------------------------------------------------------------------------------------
"""
def calculate_iou(predict_label_list, correct_label_list):
    iou_list = []

    for cor in tqdm(correct_label_list):
        time.sleep(0.000001)

        for pre in predict_label_list:
            if (cor[0] == pre[0]):
                K_area = (int(pre[4]) - int(pre[2]) + 1) * (int(pre[5]) - int(pre[3]) + 1)
                N_area = (int(cor[4]) - int(cor[2]) + 1) * (int(cor[5]) - int(cor[3]) + 1)

                iw = min(int(cor[4]), int(pre[4])) - max(int(cor[2]), int(pre[2])) + 1

                if (iw > 0):
                    ih = min(int(cor[5]), int(pre[5])) - max(int(cor[3]), int(pre[3])) + 1

                    if (ih > 0):
                        ua = (K_area + N_area) - iw*ih
                        ans = iw*ih / ua

                if (iw > 0 and ih > 0 and args.iou_threshold <= ans):
                    iou_list.append([cor[0], cor[1], pre[1], cor[6], pre[6], ans, pre[7]])

    # 昇順にソート
    iou_list.sort()

    # iou_list = [framename, cor_label, pre_label, cor_num, pre_num, iou, score]
    return iou_list


"""
---------------------------------------------------------------------------------------------------

IOUを計算（回転ver.）

---------------------------------------------------------------------------------------------------
"""
def calculate_iou_rotate(predict_label_list, correct_label_list):
    iou_list = []

    for cor in tqdm(correct_label_list):
        time.sleep(0.000001)

        # 空の配列を生成
        cor_im = np.zeros((800, 800, 3), dtype=np.uint8)

        # ポリゴンの座標を指定
        pts = np.array([[0, 0], [0, 800], [800, 800], [800, 0]])
        dst = cv2.fillPoly(cor_im, pts =[pts], color=(100, 100, 100))

        contours = np.array([[cor[2], cor[3]], [cor[4], cor[5]], [cor[8], cor[9]], [cor[6], cor[7]]])
        cor_mask = cv2.fillPoly(dst, [contours], (255, 255, 255))

        cor_area = cv2.contourArea(contours)

        for pre in predict_label_list:
            if (cor[0] == pre[0]):
                # 空の配列を生成
                pre_im = np.zeros((800, 800, 3), dtype=np.uint8)

                # ポリゴンの座標を指定
                pts = np.array([[0, 0], [0, 800], [800, 800], [800, 0]])
                dst = cv2.fillPoly(pre_im, pts =[pts], color=(0, 0, 0))

                contours = np.array([[pre[2], pre[3]], [pre[4], pre[5]], [pre[6], pre[7]], [pre[8], pre[9]]])
                pre_mask = cv2.fillPoly(dst, [contours], (255, 255, 255))

                gray_cor_mask = cv2.cvtColor(cor_mask, cv2.COLOR_BGR2GRAY)
                gray_pre_mask = cv2.cvtColor(pre_mask, cv2.COLOR_BGR2GRAY)

                pre_area = cv2.contourArea(contours)

                overlap = (gray_cor_mask == gray_pre_mask).sum()
                ans = overlap / ((cor_area + pre_area) - overlap)

                if (args.iou_threshold <= ans):
                    iou_list.append([cor[0], cor[1], pre[1], cor[10], pre[10], ans, pre[11]])

    # iou_list = [framename, cor_label, pre_label, cor_num, pre_num, iou, score]
    return iou_list


"""
---------------------------------------------------------------------------------------------------

IOUごとに評価分類（検出（正解＋推論）・不要）

---------------------------------------------------------------------------------------------------
"""
def classify_iou(iou_list):
    # ダブっている情報を削除
    for i, x in enumerate(iou_list[:]):
        for j, y in enumerate(iou_list[:]):
            if (x[0] == y[0] and x[3] != y[3] and x[4] == y[4]):
                # iouが小さい方を削除
                if (x[5] >= y[5]):
                    iou_list.remove(y)

                elif (x[5] == y[5]):
                    if (x[6] >= y[6] and y[1] != y[2]):
                        iou_list.remove(y)

                    elif (x[1] == x[2]):
                        iou_list.remove(y)

    # 不要
    unnecessary = []

    # リストから"不要"を取り出す
    for i, x in enumerate(iou_list[:]):
        for j, y in enumerate(iou_list[:]):
            if (x[0] == y[0] and x[3] == y[3] and x[4] != y[4]):
                # iouが小さい方を"不要"に
                if (x[5] >= y[5]):
                    unnecessary.append(y)
                    iou_list.remove(y)

                elif (x[5] == y[5]):
                    if (x[6] >= y[6] and y[1] != y[2]):
                        #print (x, y)
                        unnecessary.append(y)
                        iou_list.remove(y)

                    elif (x[1] == x[2]):
                        #print (x, y)
                        unnecessary.append(y)
                        iou_list.remove(y)

    # 検出（正解＋推論）
    detection = iou_list

    return detection, unnecessary


"""
---------------------------------------------------------------------------------------------------

評価ごとに数をカウント

---------------------------------------------------------------------------------------------------
"""
def count_evaluation_num(categories, pre_boxnum_list, cor_boxnum_list, detection, unnecessary):
    # 静的確保
    count_eval = [[0 for i in range(len(categories)*2 + 3)] for j in range(len(categories))]

    # 検出（正解＋推論）
    for i, x in enumerate(categories):
        for j, y in enumerate(categories):
            for obj in detection:
                if (x == obj[1] and y == obj[2]):
                    count_eval[i][1 + j] += 1

                if (x == obj[2] and y == obj[1]):
                    count_eval[i][1 + len(categories) + j] += 1

    # 不要
    for i, category in enumerate(categories):
        for obj in unnecessary:
            if (category == obj[2]):
                count_eval[i][2 + len(categories)*2] += 1

    # 過検出
    for i, box in enumerate(count_eval):
        pre_box_num  = 0

        for j in range(1 + len(categories), len(categories)*2 + 3):
            pre_box_num += count_eval[i][j]

        t = pre_boxnum_list[i][1] - pre_box_num
        count_eval[i][len(categories)*2 + 1] = t

    # 未検出
    for i, box in enumerate(count_eval):
        cor_box_num  = 0

        for j in range(1, len(categories) + 1):
            cor_box_num += count_eval[i][j]

        t = cor_boxnum_list[i][1] - cor_box_num
        count_eval[i][0] = t

    # count_eval = [未検出, 正解検出(1), ・・・, 正解検出(n), 推論検出(1), ・・・, 推論検出(n), 過検出, 不要]
    return count_eval


"""
---------------------------------------------------------------------------------------------------

平均IOUを計算

---------------------------------------------------------------------------------------------------
"""
def calculate_averageiou(categories, detection):
    avg_iou = []

    for i, category in enumerate(categories):
        div = 0
        sum_iou = 0

        for dct in detection:
            if (dct[1] == dct[2] and category == dct[1]):
                div += 1
                sum_iou += dct[5]

        # 平均を算出
        if (div == 0):
            avg = 0

        else:
            avg = '{:.2f}'.format(sum_iou/div)

        avg_iou.append(avg)

    return avg_iou


"""
---------------------------------------------------------------------------------------------------

データをエクセルに出力

---------------------------------------------------------------------------------------------------
"""
def output_excel(CLASSES, count_eval, avg_iou, detection, unnecessary, pre_label):
    cls = len(CLASSES)

    # エクセル作成
    wb = px.Workbook()

    # シート作成
    ws_1 = wb.active
    ws_1 = wb.create_sheet(title='推論結果')

    # セル結合
    ws_1.merge_cells(start_row = 1, start_column = 1,         end_row = 3, end_column = 1)
    ws_1.merge_cells(start_row = 1, start_column = 2,         end_row = 1, end_column = 2 + cls)
    ws_1.merge_cells(start_row = 2, start_column = 2,         end_row = 3, end_column = 2)
    ws_1.merge_cells(start_row = 1, start_column = 3 + cls,   end_row = 1, end_column = 4 + cls*2)
    ws_1.merge_cells(start_row = 2, start_column = 3,         end_row = 2, end_column = 2 + cls)
    ws_1.merge_cells(start_row = 2, start_column = 3 + cls,   end_row = 2, end_column = 2 + cls*2)
    ws_1.merge_cells(start_row = 2, start_column = 3 + cls*2, end_row = 3, end_column = 3 + cls*2)
    ws_1.merge_cells(start_row = 2, start_column = 4 + cls*2, end_row = 3, end_column = 4 + cls*2)
    ws_1.merge_cells(start_row = 1, start_column = 5 + cls*2, end_row = 3, end_column = 5 + cls*2)
    ws_1.merge_cells(start_row = 1, start_column = 6 + cls*2, end_row = 3, end_column = 6 + cls*2)
    ws_1.merge_cells(start_row = 1, start_column = 7 + cls*2, end_row = 3, end_column = 7 + cls*2)
    ws_1.merge_cells(start_row = 1, start_column = 8 + cls*2, end_row = 3, end_column = 8 + cls*2)

    # ヘッダーを記述
    ws_1.cell(row = 1, column = 1).value         = 'ラベル名'
    ws_1.cell(row = 1, column = 2).value         = '正解ボックス'
    ws_1.cell(row = 1, column = 3 + cls).value   = '推論ボックス'
    ws_1.cell(row = 2, column = 2).value         = '未検出'
    ws_1.cell(row = 2, column = 3).value         = '検出'
    ws_1.cell(row = 2, column = 3 + cls).value   = '検出'
    ws_1.cell(row = 2, column = 3 + cls*2).value = '過検出'
    ws_1.cell(row = 2, column = 4 + cls*2).value = '不要'
    ws_1.cell(row = 1, column = 5 + cls*2).value = '平均IOU'
    ws_1.cell(row = 1, column = 6 + cls*2).value = '再現率(Recall)'
    ws_1.cell(row = 1, column = 7 + cls*2).value = '適合率(Precision)'
    ws_1.cell(row = 1, column = 8 + cls*2).value = 'F値(F-measure)'

    # ラベル名を記述
    for i, label in enumerate(CLASSES):
        ws_1.cell(row = 4 + i, column = 1).value           = label
        ws_1.cell(row = 3,     column = 3 + i).value       = label
        ws_1.cell(row = 3,     column = 3 + cls + i).value = label

    # 評価データを記述
    for i, box in enumerate(count_eval):
        for j, num in enumerate(box):
                ws_1.cell(row = 4 + i, column = 2 + j).value = num

    # 平均IOUを記述
    for i, iou in enumerate(avg_iou):
        ws_1.cell(row = 4 + i, column = 5 + cls*2).value = iou

    # アルファベットのリストを生成
    alphabet_list = []
    if ((105 + cls*2) <= 123):
        for i in range(97, 105 + cls*2):
            alphabet_list.append(chr(i))

    else:
        for i in range(97, 123):
            alphabet_list.append(chr(i))

        cnt = 0
        end = 105 + cls*2 - 123

        alphabet_list_other = []

        for i in alphabet_list:
            for j in alphabet_list:
                if (cnt == end):
                    break
                alphabet_list_other.append(i + j)

                cnt += 1

        alphabet_list = alphabet_list + alphabet_list_other

    # セル幅を広めに修正
    for i, c in enumerate(alphabet_list):
        ws_1.column_dimensions[c].width = 14

    # 文字位置を変更
    for i in range(1, 9 + cls*2):
        for j in range(1, cls + 4):
            ws_1.cell(row = j, column = i).alignment = \
                    Alignment(horizontal = 'center', vertical = 'center')

    # 枠線を生成
    for i in range(1, 9 + cls*2):
        for j in range(1, cls + 4):
            ws_1.cell(row = j, column = i).border = \
                borders.Border(top    = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               left   = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               right  = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               bottom = borders.Side(style = borders.BORDER_THIN, color = '000000'))

    # フォントを変更
    font = Font(name = 'メイリオ')

    for i in range(1, 9 + cls*2):
        for j in range(1, cls + 4):
            ws_1.cell(row = j, column = i).font = font

    # シート作成
    ws_2 = wb.create_sheet(title='IOU詳細')

    # セル結合
    ws_2.merge_cells(start_row = 1, start_column = 7,  end_row = 2, end_column = 7)
    ws_2.merge_cells(start_row = 1, start_column = 8,  end_row = 1, end_column = 10)
    ws_2.merge_cells(start_row = 1, start_column = 11, end_row = 1, end_column = 13)

    # ヘッダーを記述
    ws_2.cell(row = 1, column = 1).value  = '画像'
    ws_2.cell(row = 1, column = 2).value  = '正解ラベル'
    ws_2.cell(row = 1, column = 3).value  = '推論ラベル'
    ws_2.cell(row = 1, column = 4).value  = 'IOU'
    ws_2.cell(row = 1, column = 5).value  = 'スコア'

    ws_2.cell(row = 1, column = 7).value  = 'ラベル名'
    ws_2.cell(row = 1, column = 8).value  = '検出'
    ws_2.cell(row = 1, column = 11).value = '過検出'

    ws_2.cell(row = 2, column = 8).value  = '0.6未満'
    ws_2.cell(row = 2, column = 9).value  = '0.6以上0.7未満'
    ws_2.cell(row = 2, column = 10).value  = '0.7以上'
    ws_2.cell(row = 2, column = 11).value = '0.6未満'
    ws_2.cell(row = 2, column = 12).value = '0.6以上0.7未満'
    ws_2.cell(row = 2, column = 13).value = '0.7以上'

    # データを記述
    for i, data in enumerate(detection):
        ws_2.cell(row = 2 + i, column = 1).value = data[0]
        ws_2.cell(row = 2 + i, column = 2).value = data[1]
        ws_2.cell(row = 2 + i, column = 3).value = data[2]
        ws_2.cell(row = 2 + i, column = 4).value = data[5]
        ws_2.cell(row = 2 + i, column = 5).value = data[6]

    cnt = 0

    for i, x in enumerate(pre_label):
        switch = 0

        for j, y in enumerate(detection):
            if (x[0] == y[0] and x[1] == y[2] and x[6] == y[4] and x[7] == y[6]):
                switch = 1

        if (switch == 0):
            ws_2.cell(row = len(detection) + 2 + cnt, column = 1).value = x[0]
            ws_2.cell(row = len(detection) + 2 + cnt, column = 3).value = x[1]
            ws_2.cell(row = len(detection) + 2 + cnt, column = 5).value = x[7]

            cnt += 1

    iou_split = [[0 for i in range(7)] for j in range(len(CLASSES))]

    for i, label in enumerate(CLASSES):
        iou_split[i][0] = label

        for data in detection:
            if (label == data[1] and data[1] == data[2] and data[5] < 0.6):
                iou_split[i][1] += 1

            elif (label == data[1] and data[1] == data[2] and data[5] >= 0.7):
                iou_split[i][3] += 1

            elif (label == data[1] and data[1] == data[2] and data[5] >= 0.6 and data[5] < 0.7):
                iou_split[i][2] += 1

            if (label == data[1] and data[1] != data[2] and data[5] < 0.6):
                iou_split[i][4] += 1

            elif (label == data[1] and data[1] != data[2] and data[5] >= 0.7):
                iou_split[i][6] += 1

            elif (label == data[1] and data[1] != data[2] and data[5] >= 0.6 and data[5] < 0.7):
                iou_split[i][5] += 1

    # データを記述
    for i, data in enumerate(iou_split):
        ws_2.cell(row = 3 + i, column = 7).value  = data[0]
        ws_2.cell(row = 3 + i, column = 8).value  = data[1]
        ws_2.cell(row = 3 + i, column = 9).value  = data[2]
        ws_2.cell(row = 3 + i, column = 10).value = data[3]
        ws_2.cell(row = 3 + i, column = 11).value = data[4]
        ws_2.cell(row = 3 + i, column = 12).value = data[5]
        ws_2.cell(row = 3 + i, column = 13).value = data[6]

    # アルファベットのリストを生成
    alphabet_list = [chr(i) for i in range(97, 110)]

    # セル幅を広めに修正
    for i, c in enumerate(alphabet_list):
        ws_2.column_dimensions[c].width = 14

    # 文字位置を変更
    for i in range(1, 6):
        ws_2.cell(row = 1, column = i).alignment = \
            Alignment(horizontal = 'center', vertical = 'center')

    for i in range(6, 14):
        for j in range(1, cls + 3):
                ws_2.cell(row = j, column = i).alignment = \
                    Alignment(horizontal = 'center', vertical = 'center')

    # 枠線を生成
    for i in range(7, 14):
        for j in range(1, cls + 3):
            ws_2.cell(row = j, column = i).border = \
                borders.Border(top    = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               left   = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               right  = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               bottom = borders.Side(style = borders.BORDER_THIN, color = '000000'))

    # フォントを変更
    for i in range(1, 14):
        for j in range(1, len(detection) + 1):
            ws_2.cell(row = j, column = i).font = font

    # シート削除
    wb.remove(wb['Sheet'])

    # エクセル保存
    save_excel_path = os.path.join(args.pre_xml.rsplit('/', 1)[0], 'evaluation.xlsx')
    wb.save(save_excel_path)


"""
---------------------------------------------------------------------------------------------------

メイン関数

---------------------------------------------------------------------------------------------------
"""
if __name__ == '__main__':
    # 存在有無をチェック
    if not (os.path.isfile(args.category)):
        sys.exit('Not exist category text file ({})'.format(args.category))

    # カテゴリの読み込み
    with open(args.category, 'r') as f:
        categories = f.read().splitlines()

    # 存在有無をチェック
    if not (os.path.isdir(args.pre_xml)):
        sys.exit('Not exist predict xml directory ({})'.format(args.pre_xml))

    if not (os.path.isdir(args.cor_xml)):
        sys.exit('Not exist correct xml directory ({})'.format(args.cor_xml))

    pre_xml_list = []
    pre_dir_bool = True

    # XMLファイルを取得
    if (len(glob.glob(os.path.join(args.pre_xml, '*.xml'))) != 0):
        pre_xml_list = glob.glob(os.path.join(args.pre_xml, '*.xml'))
        pre_dir_bool = False

    else:
        pre_xml_list = glob.glob(os.path.join(args.pre_xml, '**', '**', '*.xml'))

    cor_xml_list = []
    cor_dir_bool = True

    # XMLファイルを取得
    if (len(glob.glob(os.path.join(args.cor_xml, '*.xml'))) != 0):
        cor_xml_list = glob.glob(os.path.join(args.cor_xml, '*.xml'))
        cor_dir_bool = False

    else:
        cor_xml_list = glob.glob(os.path.join(args.cor_xml, '**', '**', '*.xml'))

    dir_bool = True

    if not (pre_dir_bool and cor_dir_bool):
        dir_bool = False

    pre_label_list = []
    cor_label_list = []

    # XMLファイルからラベルを取得
    if (True):
        pre_label_list = get_label_rotate(pre_xml_list, categories, dirform=dir_bool, correct = False)
        cor_label_list = get_label_rotate(cor_xml_list, categories, dirform=dir_bool, correct = True)

    else:
        pre_label_list = get_label(pre_xml_list, categories, dirform=dir_bool, correct = False)
        cor_label_list = get_label(cor_xml_list, categories, dirform=dir_bool, correct = True)

    # ボックス合計数を取得
    pre_boxnum_list = get_boxnum(pre_label_list, categories)
    cor_boxnum_list = get_boxnum(cor_label_list, categories)

    # 表示
    cprint('CALCULATE IOU NOW', attrs=['bold'])

    # IOUを計算
    if (True):
        iou_list = calculate_iou_rotate(pre_label_list, cor_label_list)

    else:
        iou_list = calculate_iou(pre_label_list, cor_label_list)

    # IOUごとに評価分類（検出（正解＋推論）・不要）
    detection, unnecessary = classify_iou(iou_list)

    #評価ごとに数をカウント
    count_eval = count_evaluation_num(categories, pre_boxnum_list, cor_boxnum_list, detection, unnecessary)

    # 平均IOUを計算
    avg_iou = calculate_averageiou(categories, detection)

    # 表示
    cprint('\nEXPORT EXCEL NOW\n', attrs=['bold'])

    # データをエクセルに出力
    output_excel(categories, count_eval, avg_iou, detection, unnecessary, pre_label_list)

    # 表示
    length_list = [len(s[0]) for s in pre_boxnum_list]
    length = max(length_list)

    cprint('PREDICT BOX', 'grey', attrs=['bold'])
    for i, j in pre_boxnum_list:
        cprint('%*s %d' % (length, i, j), 'blue', attrs=['bold'])

    cprint('CORRECT BOX', 'grey', attrs=['bold'])
    for i, j in cor_boxnum_list:
        cprint('%*s %d' % (length, i, j), 'red', attrs=['bold'])

    cprint('\nSUCCESSFUL COMPLETION', attrs=['bold'])
