import xml.etree.ElementTree as ET
import os
import sys
import glob
import numpy as np
import argparse
import openpyxl as px
import scipy.stats as stats
import pandas
import matplotlib.pyplot as plt
from statistics import mean, median,variance,stdev

parser = argparse.ArgumentParser(description = 'Caluclate IOU and Detection rate')

parser.add_argument('-c', '--category', help = 'Path to category text file', type = str, default = './category2.txt')
parser.add_argument('-x', '--xml', help = 'Path to predict xml directory', type = str, default = './val')
parser.add_argument('-n', '--name', help = 'Path to predict xml directory', type = str, default = '1280')

args = parser.parse_args()


"""
---------------------------------------------------------------------------------------------------

XMLファイルからラベルを取得

---------------------------------------------------------------------------------------------------
"""
def get_label(xml_list, correct = False):
    label_listtt = []
    label_list_o = []

    for xml_path in xml_list:
        tree = ET.parse(xml_path)
        directory = xml_path.rsplit('/', 3)[1]
        framename = xml_path.rsplit('/', 3)[3].rsplit('.', 1)[0]
        dir_frame = os.path.join(directory, framename)

        label_list = []
        for object in tree.findall('object'):
            labelname = object.find('category').find('value').text
            if not (correct):
                score = object.find('score').text

            for bndbox in object.findall('bndbox'):
                xmin = bndbox.find('xmin').text
                ymin = bndbox.find('ymin').text
                xmax = bndbox.find('xmax').text
                ymax = bndbox.find('ymax').text

                xmin = int(xmin)
                ymin = int(ymin)
                xmax = int(xmax)
                ymax = int(ymax)

                if not (correct):
                    label_list.append([dir_frame, labelname, xmin, ymin, xmax, ymax, score])

                else:
                    label_list.append([dir_frame, labelname, xmin, ymin, xmax, ymax])
                    label_list_o.append([dir_frame, labelname, xmin, ymin, xmax, ymax])

        # １枚の画像に含まれるバウンディングボックス数をカウント
        label_listtt.append(label_list)
    #空白を削除
    label_listtt = [i for i in label_listtt if len(i)!=0]
    # label_list = [directory/framename, labelname, x1, y1, x2, y2, num, (score)]
    return label_listtt,label_list_o

"""
---------------------------------------------------------------------------------------------------

ボックス合計数を取得

---------------------------------------------------------------------------------------------------
"""
def get_boxnum(label_list, categories):
    boxnum_list = []

    for category in categories:
        cnt = [x[1] for x in label_list].count(category)

        boxnum_list.append([category, cnt])

    return boxnum_list

def get_box_nnn(cor_label_list):
    c = []
    d = []
    for ist in cor_label_list:
        a = len(ist)
        b = ist[0][0]
        c.append([a,b])
        d.append(a)
    return c,d
    
def create_xls(dse,m,median,variance,stdev):
    # エクセル作成
    wb = px.Workbook()
    # シート作成
    ws_1 = wb.active
    ws_1 = wb.create_sheet(title='first')
    # ヘッダーを記述
    ws_1.cell(row = 1, column = 1).value          = 'ファイル名'
    ws_1.cell(row = 1, column = 2).value          = '矩形の数'
    ws_1.cell(row = 1, column = 3).value          = '平均'
    ws_1.cell(row = 2, column = 3).value          = m
    ws_1.cell(row = 1, column = 4).value          = '中央値'
    ws_1.cell(row = 2, column = 4).value          = median
    ws_1.cell(row = 1, column = 5).value          = '分散'
    ws_1.cell(row = 2, column = 5).value          = variance
    ws_1.cell(row = 1, column = 6).value          = '標準偏差'
    ws_1.cell(row = 2, column = 6).value          = stdev

    # 評価データを記述
    for i, box in enumerate(dse):
        ws_1.cell(row = 2 + i, column = 1).value = box[1]
        ws_1.cell(row = 2 + i, column = 2).value = box[0]
    # シート削除
    wb.remove(wb['Sheet'])
    # エクセル保存
    save_excel_path = os.path.join(args.xml,args.name+'count.xlsx')
    wb.save(save_excel_path)

def culcu(data):
    m = mean(data)
    median = np.median(data)
    variance = np.var(data)
    stdev = np.std(data)
    
    #stats.probplot(gte, dist="norm", plot=plt)
    plt.hist(gte, bins=100)
    plt.show()
    return m,median,variance,stdev


"""
---------------------------------------------------------------------------------------------------

メイン関数

---------------------------------------------------------------------------------------------------
"""
if __name__ == '__main__':
    # 存在有無をチェック
    if not (os.path.isfile(args.category)):
        sys.exit('Not exist category text file ({})'.format(args.category))

    # カテゴリの読み込み
    with open(args.category, 'r') as f:
        categories = f.read().splitlines()

    # 存在有無をチェック
    if not (os.path.isdir(args.xml)):
        sys.exit('Not exist xml file path ({})'.format(args.xml))

    # XMLファイルを取得
    cor_xml_list = glob.glob(os.path.join(args.xml,'**','supervised','*.xml'))

    # XMLファイルからラベルを取得
    cor_xml_list = sorted(cor_xml_list)
    cor_label_list,label_list_o = get_label(cor_xml_list, correct = True)
    #１枚の矩形の数
    dse,gte = get_box_nnn(cor_label_list)
    m,median,variance,stdev = culcu(gte)

    # ボックス合計数を取得
    #pre_boxnum_list = get_boxnum(pre_label_list, categories)
    cor_boxnum_list = get_boxnum(label_list_o, categories)
    create_xls(dse,m,median,variance,stdev)
    print(cor_boxnum_list)
