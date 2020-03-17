# -*- coding: utf-8 -*-

import xml.etree.ElementTree as ET
import os
import sys
import glob
from PIL import Image
import openpyxl as px
from openpyxl.styles import Alignment, borders
from openpyxl.styles.fonts import Font
import argparse


parser = argparse.ArgumentParser(description = 'Caluclate IOU and Detection rate')

parser.add_argument('-c', '--category', help = 'Path to category text file', type = str, default = './category.txt')
parser.add_argument('-px', '--pre_xml', help = 'Path to predict xml directory', type = str, default = './output-nms/presfam-320marge/draw-nms-0.6')
parser.add_argument('-cx', '--cor_xml', help = 'Path to correct xml directory', type = str, default = '/home/hiroshi/Dataset/folda-change/20191011矩形最終/remove-edge-val')
parser.add_argument('-i', '--iou_threshold', help = 'Threshold of IoU', type = float, default = 0.3)

args = parser.parse_args()


"""
---------------------------------------------------------------------------------------------------

XMLファイルからラベルを取得

---------------------------------------------------------------------------------------------------
"""
def get_label(xml_list, correct = False):
    label_list = []

    for xml_path in xml_list:
        tree = ET.parse(xml_path)
        directory = xml_path.rsplit('/', 3)[1]
        framename = xml_path.rsplit('/', 3)[3].rsplit('.', 1)[0]
        dir_frame = os.path.join(directory, framename)

        num = 0

        for object in tree.findall('object'):
            labelname = object.find('category').find('value').text

            if not (correct):
                score = object.find('score').text

            for bndbox in object.findall('bndbox'):
                xmin = bndbox.find('xmin').text
                ymin = bndbox.find('ymin').text
                xmax = bndbox.find('xmax').text
                ymax = bndbox.find('ymax').text

                xmin = int(xmin)
                ymin = int(ymin)
                xmax = int(xmax)
                ymax = int(ymax)

                if not (correct):
                    label_list.append([dir_frame, labelname, xmin, ymin, xmax, ymax, num, score])

                else:
                    label_list.append([dir_frame, labelname, xmin, ymin, xmax, ymax, num])

                # １枚の画像に含まれるバウンディングボックス数をカウント
                num += 1

    # label_list = [directory/framename, labelname, x1, y1, x2, y2, num, (score)]
    return label_list


"""
---------------------------------------------------------------------------------------------------

ボックス合計数を取得

---------------------------------------------------------------------------------------------------
"""
def get_boxnum(label_list, categories):
    boxnum_list = []

    for category in categories:
        cnt = [x[1] for x in label_list].count(category)

        boxnum_list.append([category, cnt])

    return boxnum_list


"""
---------------------------------------------------------------------------------------------------

IOUを計算

---------------------------------------------------------------------------------------------------
"""
def calculate_iou(predict_label_list, correct_label_list):
    iou_list = []

    for cor in correct_label_list:
        for pre in predict_label_list:
            if (cor[0] == pre[0]):
                K_area = (int(pre[4]) - int(pre[2]) + 1) * (int(pre[5]) - int(pre[3]) + 1)
                N_area = (int(cor[4]) - int(cor[2]) + 1) * (int(cor[5]) - int(cor[3]) + 1)

                iw = min(int(cor[4]), int(pre[4])) - max(int(cor[2]), int(pre[2])) + 1

                if (iw > 0):
                    ih = min(int(cor[5]), int(pre[5])) - max(int(cor[3]), int(pre[3])) + 1

                    if (ih > 0):
                        ua = (K_area + N_area) - iw*ih
                        ans = iw*ih / ua

                if (iw > 0 and ih > 0 and args.iou_threshold <= ans):
                    iou_list.append([cor[0], cor[1], pre[1], cor[6], pre[6], ans, pre[7]])

    # 昇順にソート
    iou_list.sort()

    # iou_list = [inference/directory/framename, cor_label, pre_label, cor_num, pre_num, iou, score]
    return iou_list


"""
---------------------------------------------------------------------------------------------------

IOUごとに評価分類（検出（正解＋推論）・不要）

---------------------------------------------------------------------------------------------------
"""
def classify_iou(iou_list):
    # ダブっている情報を削除
    for i, x in enumerate(iou_list[:]):
        for j, y in enumerate(iou_list[:]):
            if (x[0] == y[0] and x[3] != y[3] and x[4] == y[4]):
                # iouが小さい方を削除
                if (x[5] >= y[5]):
                    iou_list.remove(y)

    # 不要
    unnecessary = []

    # リストから"不要"を取り出す
    for i, x in enumerate(iou_list[:]):
        for j, y in enumerate(iou_list[:]):
            if (x[0] == y[0] and x[3] == y[3] and x[4] != y[4]):
                # iouが小さい方を"不要"に
                if (x[5] >= y[5]):
                    unnecessary.append(y)
                    iou_list.remove(y)

    # 検出（正解＋推論）
    detection = iou_list

    return detection, unnecessary


"""
---------------------------------------------------------------------------------------------------

評価ごとに数をカウント

---------------------------------------------------------------------------------------------------
"""
def count_evaluation_num(categories, pre_boxnum_list, cor_boxnum_list, detection, unnecessary):
    # 静的確保
    count_eval = [[0 for i in range(len(categories)*2 + 3)] for j in range(len(categories))]

    # 検出（正解＋推論）
    for i, x in enumerate(categories):
        for j, y in enumerate(categories):
            for obj in detection:
                if (x == obj[1] and y == obj[2]):
                    count_eval[i][1 + j] += 1

                if (x == obj[2] and y == obj[1]):
                    count_eval[i][1 + len(categories) + j] += 1

    # 不要
    for i, category in enumerate(categories):
        for obj in unnecessary:
            if (category == obj[2]):
                count_eval[i][2 + len(categories)*2] += 1

    # 過検出
    for i, box in enumerate(count_eval):
        pre_box_num  = 0

        for j in range(1 + len(categories), len(categories)*2 + 3):
            pre_box_num += count_eval[i][j]

        t = pre_boxnum_list[i][1] - pre_box_num
        count_eval[i][len(categories)*2 + 1] = t

    # 未検出
    for i, box in enumerate(count_eval):
        cor_box_num  = 0

        for j in range(1, len(categories) + 1):
            cor_box_num += count_eval[i][j]

        t = cor_boxnum_list[i][1] - cor_box_num
        count_eval[i][0] = t

    # count_eval = [未検出, 正解検出(1), ・・・, 正解検出(n), 推論検出(1), ・・・, 推論検出(n), 過検出, 不要]
    return count_eval


"""
---------------------------------------------------------------------------------------------------

平均IOUを計算

---------------------------------------------------------------------------------------------------
"""
def calculate_averageiou(categories, detection):
    avg_iou = []

    for i, category in enumerate(categories):
        div = 0
        sum_iou = 0

        for dct in detection:
            if (dct[1] == dct[2] and category == dct[1]):
                div += 1
                sum_iou += dct[5]

        # 平均を算出
        if (div == 0):
            avg = 0

        else:
            avg = '{:.2f}'.format(sum_iou/div)

        avg_iou.append(avg)

    return avg_iou

"""
---------------------------------------------------------------------------------------------------

再現率を計算

---------------------------------------------------------------------------------------------------
"""
def calculate_recall(count_eval):
    recalls= []
    size = len(categories)
    
    for i, box in enumerate(count_eval):
        Sum = 0
        for j, num in enumerate(box):
            if j <= size:
                Sum += int(num)
        if Sum != 0:
            recall = int(box[1+i]) / Sum 
            recalls.append(recall)
        else:
            recalls.append(0)

    return recalls

"""
---------------------------------------------------------------------------------------------------

適合率を計算

---------------------------------------------------------------------------------------------------
"""
def calculate_precisions(count_eval):
    precisions= []
    size = len(categories)
    
    for i, box in enumerate(count_eval):
        Sum = 0
        for j, num in enumerate(box):
            if  size < j < (size+1) * 2:
                Sum += int(num)
        if Sum != 0:
            precision = int(box[size+1+i]) / Sum 
            precisions.append(precision)
        else:
            precisions.append(0)

    return precisions
"""
---------------------------------------------------------------------------------------------------

データをエクセルに出力

---------------------------------------------------------------------------------------------------
"""
def output_excel(categories, count_eval, avg_iou, detection, unnecessary, pre_label_list, recalls, precisions):
    size = len(categories)

    # エクセル作成
    wb = px.Workbook()

    # シート作成
    ws_1 = wb.active
    ws_1 = wb.create_sheet(title='推論結果')

    # セル結合
    ws_1.merge_cells(start_row = 1, start_column = 1,          end_row = 3, end_column = 1)
    ws_1.merge_cells(start_row = 1, start_column = 2,          end_row = 1, end_column = 2 + size)
    ws_1.merge_cells(start_row = 2, start_column = 2,          end_row = 3, end_column = 2)
    ws_1.merge_cells(start_row = 1, start_column = 3 + size,   end_row = 1, end_column = 4 + size*2)
    ws_1.merge_cells(start_row = 2, start_column = 3,          end_row = 2, end_column = 2 + size)
    ws_1.merge_cells(start_row = 2, start_column = 3 + size,   end_row = 2, end_column = 2 + size*2)
    ws_1.merge_cells(start_row = 2, start_column = 3 + size*2, end_row = 3, end_column = 3 + size*2)
    ws_1.merge_cells(start_row = 2, start_column = 4 + size*2, end_row = 3, end_column = 4 + size*2)
    ws_1.merge_cells(start_row = 1, start_column = 5 + size*2, end_row = 3, end_column = 5 + size*2)
    ws_1.merge_cells(start_row = 1, start_column = 6 + size*2, end_row = 3, end_column = 6 + size*2)
    ws_1.merge_cells(start_row = 1, start_column = 7 + size*2, end_row = 3, end_column = 7 + size*2)
    ws_1.merge_cells(start_row = 1, start_column = 8 + size*2, end_row = 3, end_column = 8 + size*2)

    # ヘッダーを記述
    ws_1.cell(row = 1, column = 1).value          = 'ラベル名'
    ws_1.cell(row = 1, column = 2).value          = '正解ボックス'
    ws_1.cell(row = 1, column = 3 + size).value   = '推論ボックス'
    ws_1.cell(row = 2, column = 2).value          = '未検出'
    ws_1.cell(row = 2, column = 3).value          = '検出'
    ws_1.cell(row = 2, column = 3 + size).value   = '検出'
    ws_1.cell(row = 2, column = 3 + size*2).value = '過検出'
    ws_1.cell(row = 2, column = 4 + size*2).value = '不要'
    ws_1.cell(row = 1, column = 5 + size*2).value = '平均IOU'
    ws_1.cell(row = 1, column = 6 + size*2).value = '再現率(Recall)'
    ws_1.cell(row = 1, column = 7 + size*2).value = '適合率(Precision)'
    ws_1.cell(row = 1, column = 8 + size*2).value = 'F値(F-measure)'

    # ラベル名を記述
    for i, label in enumerate(categories):
        ws_1.cell(row = 4 + i, column = 1).value            = label
        ws_1.cell(row = 3,     column = 3 + i).value        = label
        ws_1.cell(row = 3,     column = 3 + size + i).value = label

    # 評価データを記述
    for i, box in enumerate(count_eval):
        for j, num in enumerate(box):
                ws_1.cell(row = 4 + i, column = 2 + j).value = num

    # 平均IOUを記述
    for i, iou in enumerate(avg_iou):
        ws_1.cell(row = 4 + i, column = 5 + size*2).value = iou

    #再現率、適合率、F値を記述
    for i in range(len(recalls)):
        ws_1.cell(row = 4 + i, column = 6 + size*2).value = recalls[i]
        ws_1.cell(row = 4 + i, column = 7 + size*2).value = precisions[i]
        if  recalls[i] != 0 and precisions[i] != 0:
            fmeasure = 2 * recalls[i] * precisions[i] / ( recalls[i] + precisions[i] )
        else:
            fmeasure = 0
        ws_1.cell(row = 4 + i, column = 8 + size*2).value = fmeasure

    # アルファベットのリストを生成
    alphabet_list = [chr(i) for i in range(97, 105 + size*2)]

    # セル幅を広めに修正
   # for i, c in enumerate(alphabet_list):
    #    ws_1.column_dimensions[c].width = 14

    # 文字位置を変更
    for i in range(1, 9 + size*2):
        for j in range(1, size + 4):
            ws_1.cell(row = j, column = i).alignment = \
                    Alignment(horizontal = 'center', vertical = 'center')

    # 枠線を生成
    for i in range(1, 9 + size*2):
        for j in range(1, size + 4):
            ws_1.cell(row = j, column = i).border = \
                borders.Border(top    = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               left   = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               right  = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               bottom = borders.Side(style = borders.BORDER_THIN, color = '000000'))

    # フォントを変更
    font = Font(name = 'メイリオ')

    for i in range(1, 9 + size*2):
        for j in range(1, size + 4):
            ws_1.cell(row = j, column = i).font = font

    # シート作成
    ws_2 = wb.create_sheet(title='IOU詳細')

    # セル結合
    ws_2.merge_cells(start_row = 1, start_column = 7, end_row = 2, end_column = 7)
    ws_2.merge_cells(start_row = 1, start_column = 8, end_row = 1, end_column = 10)
    ws_2.merge_cells(start_row = 1, start_column = 11, end_row = 1, end_column = 13)

    # ヘッダーを記述
    ws_2.cell(row = 1, column = 1).value  = '画像'
    ws_2.cell(row = 1, column = 2).value  = '正解ラベル'
    ws_2.cell(row = 1, column = 3).value  = '推論ラベル'
    ws_2.cell(row = 1, column = 4).value  = 'IOU'
    ws_2.cell(row = 1, column = 5).value  = 'スコア'

    ws_2.cell(row = 1, column = 7).value  = 'ラベル名'
    ws_2.cell(row = 1, column = 8).value  = '検出'
    ws_2.cell(row = 1, column = 11).value = '過検出'

    ws_2.cell(row = 2, column = 8).value  = '0.6未満'
    ws_2.cell(row = 2, column = 9).value  = '0.6以上0.7未満'
    ws_2.cell(row = 2, column = 10).value = '0.7以上'
    ws_2.cell(row = 2, column = 11).value = '0.6未満'
    ws_2.cell(row = 2, column = 12).value = '0.6以上0.7未満'
    ws_2.cell(row = 2, column = 13).value = '0.7以上'

    # データを記述
    for i, data in enumerate(detection):
        ws_2.cell(row = 2 + i, column = 1).value = data[0]
        ws_2.cell(row = 2 + i, column = 2).value = data[1]
        ws_2.cell(row = 2 + i, column = 3).value = data[2]
        ws_2.cell(row = 2 + i, column = 4).value = data[5]
        ws_2.cell(row = 2 + i, column = 5).value = data[6]

    cnt = 0

    for i, x in enumerate(pre_label_list):
        switch = 0

        for j, y in enumerate(detection):
            if (x[0] == y[0] and x[1] == y[2] and x[6] == y[4] and x[7] == y[6]):
                switch = 1

        if (switch == 0):
            ws_2.cell(row = len(detection) + 2 + cnt, column = 1).value = x[0]
            ws_2.cell(row = len(detection) + 2 + cnt, column = 3).value = x[1]
            ws_2.cell(row = len(detection) + 2 + cnt, column = 5).value = x[7]

            cnt += 1

    iou_split = [[0 for i in range(7)] for j in range(len(categories))]

    for i, label in enumerate(categories):
        iou_split[i][0] = label

        for data in detection:
            if (label == data[1] and data[1] == data[2] and data[5] < 0.6):
                iou_split[i][1] += 1

            elif (label == data[1] and data[1] == data[2] and data[5] >= 0.7):
                iou_split[i][3] += 1

            elif (label == data[1] and data[1] == data[2] and data[5] >= 0.6 and data[5] < 0.7):
                iou_split[i][2] += 1

            if (label == data[1] and data[1] != data[2] and data[5] < 0.6):
                iou_split[i][4] += 1

            elif (label == data[1] and data[1] != data[2] and data[5] >= 0.7):
                iou_split[i][6] += 1

            elif (label == data[1] and data[1] != data[2] and data[5] >= 0.6 and data[5] < 0.7):
                iou_split[i][5] += 1

    # データを記述
    for i, data in enumerate(iou_split):
        ws_2.cell(row = 3 + i, column = 7).value   = data[0]
        ws_2.cell(row = 3 + i, column = 8).value   = data[1]
        ws_2.cell(row = 3 + i, column = 9).value   = data[2]
        ws_2.cell(row = 3 + i, column = 10).value  = data[3]
        ws_2.cell(row = 3 + i, column = 11).value  = data[4]
        ws_2.cell(row = 3 + i, column = 12).value  = data[5]
        ws_2.cell(row = 3 + i, column = 13).value  = data[6]

    # アルファベットのリストを生成
    alphabet_list = [chr(i) for i in range(97, 110)]

    # セル幅を広めに修正
    for i, c in enumerate(alphabet_list):
        ws_2.column_dimensions[c].width = 14

    # 文字位置を変更
    for i in range(1, 6):
        ws_2.cell(row = 1, column = i).alignment = \
            Alignment(horizontal = 'center', vertical = 'center')

    for i in range(6, 14):
        for j in range(1, size + 3):
                ws_2.cell(row = j, column = i).alignment = \
                    Alignment(horizontal = 'center', vertical = 'center')

    # 枠線を生成
    for i in range(7, 14):
        for j in range(1, size + 3):
            ws_2.cell(row = j, column = i).border = \
                borders.Border(top    = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               left   = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               right  = borders.Side(style = borders.BORDER_THIN, color = '000000'),
                               bottom = borders.Side(style = borders.BORDER_THIN, color = '000000'))

    # フォントを変更
    for i in range(1, 14):
        for j in range(1, len(detection) + 1):
            ws_2.cell(row = j, column = i).font = font

    # シート削除
    wb.remove(wb['Sheet'])

    # エクセル保存
    save_excel_path = os.path.join(args.pre_xml, 'evaluation.xlsx')
    wb.save(save_excel_path)

    print ('Successful Completion ({})'.format(save_excel_path))


"""
---------------------------------------------------------------------------------------------------

メイン関数

---------------------------------------------------------------------------------------------------
"""
if __name__ == '__main__':
    # 存在有無をチェック
    if not (os.path.isfile(args.category)):
        sys.exit('Not exist category text file ({})'.format(args.category))

    # カテゴリの読み込み
    with open(args.category, 'r') as f:
        categories = f.read().splitlines()

    # 存在有無をチェック
    if not (os.path.isdir(args.pre_xml)):
        sys.exit('Not exist xml file path ({})'.format(args.pre_xml))

    if not (os.path.isdir(args.cor_xml)):
        sys.exit('Not exist xml file path ({})'.format(args.cor_xml))

    # XMLファイルを取得
    pre_xml_list = glob.glob(os.path.join(args.pre_xml, '**','supervised','*.xml'))
    cor_xml_list = glob.glob(os.path.join(args.cor_xml, '**', 'supervised', '*.xml'))

    # XMLファイルからラベルを取得
    pre_label_list = get_label(pre_xml_list, correct = False)
    cor_label_list = get_label(cor_xml_list, correct = True)

    # ボックス合計数を取得
    pre_boxnum_list = get_boxnum(pre_label_list, categories)
    cor_boxnum_list = get_boxnum(cor_label_list, categories)

    # IOUを計算
    iou_list = calculate_iou(pre_label_list, cor_label_list)

    # IOUごとに評価分類（検出（正解＋推論）・不要）
    detection, unnecessary = classify_iou(iou_list)

    #評価ごとに数をカウント
    count_eval = count_evaluation_num(categories, pre_boxnum_list, cor_boxnum_list, detection, unnecessary)

    # 平均IOUを計算
    avg_iou = calculate_averageiou(categories, detection)

    # 再現率を計算
    recalls = calculate_recall(count_eval)

    # 適合率を計算
    precisions = calculate_precisions(count_eval)

    # データをエクセルに出力
    output_excel(categories, count_eval, avg_iou, detection, unnecessary, pre_label_list, recalls, precisions)
